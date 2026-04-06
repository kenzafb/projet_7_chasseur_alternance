"""
export_excel.py
===============
Génère un fichier Excel de suivi des candidatures spontanées
à partir de entreprises_enrichies.json.

Feuilles :
  - Candidatures  : tableau de suivi avec statuts, colonnes à remplir
  - Stats         : résumé automatique (formules Excel)
  - Source        : toutes les données brutes importées

Usage : python export_excel.py
"""

import json
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

FICHIER_ENTREE = "data/entreprises_enrichies.json"
FICHIER_SORTIE = f"candidatures_spontanees_{datetime.today().strftime('%Y%m')}.xlsx"

# ─── Couleurs ─────────────────────────────────────────────────────────────────
VIOLET_FONCE  = "4B3F72"
VIOLET_CLAIR  = "EDE9F7"
VERT          = "27AE60"
VERT_CLAIR    = "E8F8F0"
ORANGE        = "E67E22"
ORANGE_CLAIR  = "FEF0E0"
ROUGE         = "E74C3C"
ROUGE_CLAIR   = "FDEDEC"
GRIS_CLAIR    = "F5F5F5"
GRIS_TITRE    = "2C3E50"
BLANC         = "FFFFFF"
BLEU_CLAIR    = "EBF5FB"

STATUTS = ["À envoyer", "Envoyée", "Relancée", "Entretien", "Refus", "Sans réponse"]
COULEURS_STATUTS = {
    "À envoyer":   ("FFF3CD", "856404"),
    "Envoyée":     ("CCE5FF", "004085"),
    "Relancée":    ("D4EDDA", "155724"),
    "Entretien":   ("D1ECF1", "0C5460"),
    "Refus":       ("F8D7DA", "721C24"),
    "Sans réponse":("E2E3E5", "383D41"),
}

# ─── Styles helpers ───────────────────────────────────────────────────────────

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, bold=False, bg=None, fg="000000",
             size=11, align="left", italic=False, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg, size=size, italic=italic, name="Arial")
    cell.alignment = center() if align == "center" else left()
    if bg:
        cell.fill = fill(bg)
    if border:
        cell.border = thin_border()
    return cell


# ─── Feuille Candidatures ─────────────────────────────────────────────────────

def creer_feuille_candidatures(wb, entreprises):
    ws = wb.active
    ws.title = "Candidatures"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"  # Figer les 2 premières lignes

    # Titre principal
    ws.merge_cells("A1:N1")
    titre = ws["A1"]
    titre.value = f"🎯 Suivi Candidatures Spontanées — Kenza | DEUST IOSI DevOps | Rentrée Sept. 2026"
    titre.font = Font(bold=True, color=BLANC, size=13, name="Arial")
    titre.fill = fill(VIOLET_FONCE)
    titre.alignment = center()
    ws.row_dimensions[1].height = 35

    # En-têtes
    colonnes = [
        ("N°",            5),
        ("Entreprise",    30),
        ("Ville",         14),
        ("Code NAF",      10),
        ("Taille",        10),
        ("Site Web",      28),
        ("Email Contact", 32),
        ("Statut",        14),
        ("Date Envoi",    13),
        ("Date Relance",  13),
        ("Réponse",       14),
        ("Contact RH",    20),
        ("Notes",         30),
        ("Lettre Gemini", 14),
    ]

    for col_idx, (nom, largeur) in enumerate(colonnes, 1):
        cell = ws.cell(row=2, column=col_idx, value=nom)
        cell.font = Font(bold=True, color=BLANC, size=10, name="Arial")
        cell.fill = fill(GRIS_TITRE)
        cell.alignment = center()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = largeur

    ws.row_dimensions[2].height = 28

    # Validation liste déroulante Statut (colonne H = 8)
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(STATUTS)}"',
        allow_blank=True,
        showDropDown=False
    )
    dv.error = "Choisir un statut valide"
    dv.errorTitle = "Statut invalide"
    dv.prompt = "Choisir le statut de la candidature"
    ws.add_data_validation(dv)

    # Données
    ligne = 3
    avec_email = [e for e in entreprises if e.get("emails_trouves") or e.get("emails")]
    sans_email  = [e for e in entreprises if not e.get("emails_trouves") and not e.get("emails")]

    # On met d'abord celles avec email
    toutes = avec_email + sans_email

    for idx, e in enumerate(toutes, 1):
        emails = e.get("emails_trouves") or e.get("emails") or []
        email_principal = emails[0] if emails else ""

        bg_ligne = BLANC if idx % 2 == 0 else GRIS_CLAIR

        valeurs = [
            idx,
            e.get("nom", ""),
            e.get("ville", ""),
            e.get("code_naf", ""),
            e.get("tranche_effectif", ""),
            e.get("site_web") or e.get("url_scrapee") or "",
            email_principal,
            "À envoyer",  # Statut par défaut
            "",  # Date envoi
            "",  # Date relance
            "",  # Réponse
            "",  # Contact RH
            "",  # Notes
            "Non",  # Lettre Gemini
        ]

        for col_idx, val in enumerate(valeurs, 1):
            cell = ws.cell(row=ligne, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10, color="000000")
            cell.border = thin_border()
            cell.fill = fill(bg_ligne)
            cell.alignment = left()

            # URL cliquable
            if col_idx == 6 and val and val.startswith("http"):
                cell.hyperlink = val
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

            # Email cliquable
            if col_idx == 7 and val and "@" in val:
                cell.hyperlink = f"mailto:{val}"
                cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")

            # Colonne Statut : liste déroulante
            if col_idx == 8:
                dv.add(cell)
                cell.alignment = center()

        ws.row_dimensions[ligne].height = 22
        ligne += 1

    # Mise en forme conditionnelle légère via note visuelle
    print(f"  [Excel] {ligne - 3} lignes entreprises ajoutées")
    return ligne


# ─── Feuille Stats ─────────────────────────────────────────────────────────────

def creer_feuille_stats(wb, nb_total):
    ws = wb.create_sheet("Stats")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "📊 Tableau de Bord Candidatures"
    t.font = Font(bold=True, color=BLANC, size=13, name="Arial")
    t.fill = fill(VIOLET_FONCE)
    t.alignment = center()
    ws.row_dimensions[1].height = 32

    indicateurs = [
        ("Total entreprises",    f"=COUNTA(Candidatures!B3:B{2+nb_total})", VIOLET_CLAIR),
        ("À envoyer",            f'=COUNTIF(Candidatures!H3:H{2+nb_total},"À envoyer")', "FFF3CD"),
        ("Envoyées",             f'=COUNTIF(Candidatures!H3:H{2+nb_total},"Envoyée")', "CCE5FF"),
        ("Relancées",            f'=COUNTIF(Candidatures!H3:H{2+nb_total},"Relancée")', "D4EDDA"),
        ("Entretiens",           f'=COUNTIF(Candidatures!H3:H{2+nb_total},"Entretien")', "D1ECF1"),
        ("Refus",                f'=COUNTIF(Candidatures!H3:H{2+nb_total},"Refus")', ROUGE_CLAIR),
        ("Sans réponse",         f'=COUNTIF(Candidatures!H3:H{2+nb_total},"Sans réponse")', GRIS_CLAIR),
        ("Taux de réponse (%)",  f'=IF(COUNTIF(Candidatures!H3:H{2+nb_total},"Envoyée")=0,"—",ROUND((COUNTIF(Candidatures!H3:H{2+nb_total},"Entretien")+COUNTIF(Candidatures!H3:H{2+nb_total},"Refus"))*100/COUNTIF(Candidatures!H3:H{2+nb_total},"Envoyée"),1))', VERT_CLAIR),
    ]

    for row, (label, formule, bg) in enumerate(indicateurs, 3):
        set_cell(ws, row, 1, label, bold=True, bg=bg, size=11)
        cell_val = ws.cell(row=row, column=2, value=formule)
        cell_val.font = Font(bold=True, size=13, color=VIOLET_FONCE, name="Arial")
        cell_val.fill = fill(bg)
        cell_val.alignment = center()
        cell_val.border = thin_border()
        ws.row_dimensions[row].height = 26

    # Légende statuts
    ws.cell(row=13, column=1).value = "Légende des statuts"
    ws.cell(row=13, column=1).font = Font(bold=True, size=11, name="Arial")
    ws.row_dimensions[13].height = 22

    for r, (statut, (bg, fg)) in enumerate(COULEURS_STATUTS.items(), 14):
        c = ws.cell(row=r, column=1, value=f"  {statut}")
        c.fill = fill(bg)
        c.font = Font(color=fg, name="Arial", size=10)
        c.border = thin_border()
        c.alignment = left()
        ws.row_dimensions[r].height = 20


# ─── Feuille Source ───────────────────────────────────────────────────────────

def creer_feuille_source(wb, entreprises):
    ws = wb.create_sheet("Données brutes")
    ws.sheet_view.showGridLines = False

    headers = ["Nom", "SIRET", "SIREN", "Code NAF", "Libellé NAF",
               "Adresse", "Ville", "Dept", "Taille", "Site Web", "Emails trouvés", "Traité"]

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color=BLANC, size=10, name="Arial")
        c.fill = fill(GRIS_TITRE)
        c.alignment = center()
        c.border = thin_border()

    largeurs = [35, 16, 12, 10, 30, 35, 18, 8, 10, 35, 45, 8]
    for col, w in enumerate(largeurs, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    for row, e in enumerate(entreprises, 2):
        emails = e.get("emails_trouves") or e.get("emails") or []
        valeurs = [
            e.get("nom", ""),
            e.get("siret", ""),
            e.get("siren", ""),
            e.get("code_naf", ""),
            e.get("libelle_naf", ""),
            e.get("adresse", ""),
            e.get("ville", ""),
            e.get("departement", ""),
            e.get("tranche_effectif", ""),
            e.get("site_web") or e.get("url_scrapee") or "",
            " | ".join(emails),
            "✅" if e.get("traite") else "⏳",
        ]
        bg = BLANC if row % 2 == 0 else GRIS_CLAIR
        for col, val in enumerate(valeurs, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = Font(name="Arial", size=9)
            c.fill = fill(bg)
            c.border = thin_border()
            c.alignment = left()
        ws.row_dimensions[row].height = 18


# ─── Point d'entrée ───────────────────────────────────────────────────────────

def main():
    print("=" * 55)
    print("  Chasseur d'Alternance — Export Excel")
    print("=" * 55)

    # Charger les données (enrichies si dispo, sinon raw)
    fichier = FICHIER_ENTREE
    if not os.path.exists(fichier):
        fichier = "entreprises_raw.json"
    if not os.path.exists(fichier):
        print("[ERREUR] Aucun fichier JSON trouvé. Lance d'abord fetch_entreprises.py")
        return

    with open(fichier, "r", encoding="utf-8") as f:
        entreprises = json.load(f)

    print(f"[Chargement] {len(entreprises)} entreprises depuis {fichier}")

    # Trier : avec email d'abord, puis par ville
    entreprises.sort(key=lambda e: (
        0 if (e.get("emails_trouves") or e.get("emails")) else 1,
        e.get("ville", "")
    ))

    wb = Workbook()

    print("[Excel] Création feuille Candidatures...")
    nb = creer_feuille_candidatures(wb, entreprises)

    print("[Excel] Création feuille Stats...")
    creer_feuille_stats(wb, len(entreprises))

    print("[Excel] Création feuille Données brutes...")
    creer_feuille_source(wb, entreprises)

    # Sauvegarder
    output_path = FICHIER_SORTIE
    wb.save(output_path)

    avec_email = sum(1 for e in entreprises if e.get("emails_trouves") or e.get("emails"))
    print("\n" + "=" * 55)
    print(f"  Entreprises exportées : {len(entreprises)}")
    print(f"  Avec email            : {avec_email}")
    print(f"  Fichier Excel         : {FICHIER_SORTIE}")
    print("=" * 55)


if __name__ == "__main__":
    main()
